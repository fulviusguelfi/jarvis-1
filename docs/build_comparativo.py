#!/usr/bin/env python3
"""Gera comparativo_modelos_LLM.xlsx — comparacao de modelos para o Jarvis-1.

Dados de benchmark: fontes oficiais (HF model cards, qwen.ai, papers) coletadas em jun/2026.
Colunas de hardware (footprint/velocidade): calculadas para RX 580 8GB + 16GB RAM + Ryzen 5 4500.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"

# Cores
C_HDR   = "1F3864"   # azul escuro header
C_BASE  = "C00000"   # vermelho (modelo atual / pesado)
C_PART  = "FFE699"   # amarelo (parcial)
C_GOOD  = "C6EFCE"   # verde (recomendado / cabe)
C_GOODT = "006100"   # verde texto
C_BADT  = "9C0006"   # vermelho texto
C_GRP   = "D9E1F2"   # azul claro faixa de grupo
C_ALT   = "F2F2F2"   # zebra

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---------------------------------------------------------------------------
# DADOS  (n/d = nao coletado de fonte primaria nesta pesquisa)
# Ordem: por params totais desc. Modelo atual no topo.
# campos: modelo, dev, arq, ptot, pativ, ctx, lic, q4gb, vram, ramfree, vel,
#         mmlupro, gpqa, ifeval, bfcl, ptbr, tools, tier, veredito, fonte
# ---------------------------------------------------------------------------
M = [
 ["Qwen3.6-35B-A3B  (ATUAL)","Qwen","Hibrido SSM+MoE (DeltaNet 3:1)",35,3.0,"262K (->1M)","Apache-2.0",19.7,
  "Nao","~0 livre (ocupa tudo)","~5-8",
  "n/d","n/d","n/d","n/d","Bom (201 idiomas)","Nativo (forte)","Muito Alto",
  "Capaz, mas ocupa RAM+VRAM inteiras -> derrota o proposito de assistente local","qwen.ai / HF HauhauCS"],

 ["Qwen3-30B-A3B-Instruct-2507","Qwen","MoE (3B ativos)",30,3.0,"262K","Apache-2.0",18.0,
  "Nao","~0 livre","~5-9",
  69.1,54.8,83.7,58.6,"Bom (119 idiomas)","Nativo (forte)","Alto",
  "Mesmo problema de memoria do 35B; o 4B-2507 empata com ele","HF Qwen3-4B-2507 card"],

 ["Gemma 3 27B","Google","Densa",27,27,"128K","Gemma (restritiva)",16.0,
  "Nao","~0 livre","~2-4",
  "n/d (52.2 CoT)","24.3","n/d","n/d","Excelente (140 idiomas)","Suportado","Alto",
  "Otimo multilingue, mas pesado demais p/ este hardware","HF google/gemma-3"],

 ["Qwen3-32B","Qwen","Densa",32,32,"128K","Apache-2.0",20.0,
  "Nao","~0 livre","~2-3",
  "n/d","n/d","n/d","n/d","Bom","Nativo (forte)","Alto",
  "Denso de 32B: lentissimo e ocupa tudo neste hardware","HF Qwen3 / insiderllm"],

 ["Phi-4 (14B)","Microsoft","Densa",14,14,"16K","MIT",9.0,
  "Parcial","~6-7 GB","~8-12",
  "n/d","n/d","n/d","n/d","Razoavel (EN-centrico)","Suportado","Alto (STEM)",
  "Forte em raciocinio/STEM; contexto curto (16K) e offload parcial","HF microsoft/phi-4"],

 ["Qwen3-14B","Qwen","Densa",14,14,"128K","Apache-2.0",9.0,
  "Parcial","~6-7 GB","~8-12",
  "n/d","n/d","n/d","n/d","Bom","Nativo (forte)","Alto",
  "Bom equilibrio qualidade, mas nao cabe 100% na VRAM (offload parcial)","HF Qwen3 / insiderllm"],

 ["Gemma 3 12B","Google","Densa",12,12,"128K","Gemma (restritiva)",7.3,
  "Parcial","~7-8 GB","~12-18",
  "n/d","25.4","n/d","n/d","Excelente (140 idiomas)","Suportado","Medio-Alto",
  "Multilingue forte; cabe quase 100% em VRAM com contexto modesto","HF google/gemma-3"],

 ["Mistral Small 3 (24B)","Mistral","Densa",24,24,"32K","Apache-2.0",14.0,
  "Nao","~0-2 GB","~3-5",
  "n/d","n/d","n/d","n/d","Bom (idiomas EU + PT)","Nativo","Alto",
  "Forte, mas 24B pesa demais aqui","Mistral AI"],

 ["Granite 3.3 8B","IBM","Densa",8,8,"128K","Apache-2.0",4.9,
  "Sim","~9-10 GB","~18-28",
  "n/d","n/d","n/d","n/d","Razoavel (12 idiomas)","Nativo (projetado p/ tools)","Medio",
  "Cabe 100% na VRAM; bom p/ tool calling empresarial","IBM Granite"],

 ["Llama 3.1 8B Instruct","Meta","Densa",8,8,"128K","Llama 3.1 (custom)",4.9,
  "Sim","~9-10 GB","~18-28",
  "n/d","32.8","80.4","76.1","Bom (8 idiomas, inc. PT)","Nativo (BFCL 76)","Medio",
  "Cabe 100% na VRAM; tool calling forte; PT-BR oficial","HF meta-llama / Llama3 paper"],

 ["Qwen3-8B","Qwen","Densa",8,8,"128K","Apache-2.0",4.9,
  "Sim","~9-10 GB","~18-28",
  "n/d","n/d","n/d","n/d","Bom (119 idiomas)","Nativo (forte)","Medio-Alto",
  "DROP-IN no codigo atual; cabe 100% na VRAM; ja foi usado no projeto","HF Qwen3 / insiderllm"],

 ["Ministral 8B (Ministral 3)","Mistral","Densa",8,8,"128K","Mistral Research/Comm.",4.9,
  "Sim","~9-10 GB","~18-28",
  "n/d","n/d","n/d","n/d","Bom (idiomas EU + PT)","Nativo","Medio",
  "Cabe 100% na VRAM; eficiente; supera Gemma 12B em parte dos evals","Ministral 3 paper (arXiv)"],

 ["Gemma 3 4B","Google","Densa",4,4,"128K","Gemma (restritiva)",2.5,
  "Sim","~10-11 GB","~30-45",
  "n/d (Glob-MMLU 57)","15.0","n/d","n/d","Excelente (140 idiomas)","Suportado","Medio",
  "MELHOR multilingue leve; cabe folgado; tool calling menos maduro","HF google/gemma-3"],

 ["Qwen3-4B-Instruct-2507  *","Qwen","Densa",4,4,"262K","Apache-2.0",2.5,
  "Sim","~10-11 GB livre","~30-45",
  69.6,62.0,83.4,61.9,"Bom (119 idiomas)","Nativo (BFCL 61.9)","Medio-Alto",
  "ESCOLHA TOP: empata/supera o 30B-A3B; DROP-IN; ja baixado; deixa a maquina livre","HF Qwen/Qwen3-4B-Instruct-2507"],

 ["Phi-4-mini (3.8B)","Microsoft","Densa",3.8,3.8,"128K","MIT",2.5,
  "Sim","~10-11 GB","~30-45",
  52.8,25.2,"n/d","n/d","Razoavel","Suportado","Medio",
  "Forte em mat/STEM p/ o tamanho; PT-BR mediano","HF microsoft/Phi-4-mini-instruct"],

 ["Llama 3.2 3B Instruct","Meta","Densa",3.0,3.0,"128K","Llama 3.2 (custom)",2.0,
  "Sim","~11-12 GB","~35-50",
  39.2,32.8,77.4,67.0,"Bom (8 idiomas, inc. PT)","Nativo (BFCL 67)","Medio-Basico",
  "Leve, tool calling decente; conhecimento limitado (MMLU-Pro 39)","HF meta-llama/Llama-3.2-3B"],

 ["Qwen3-1.7B","Qwen","Densa",1.7,1.7,"32K","Apache-2.0",1.1,
  "Sim","~12-13 GB","~50-70",
  "n/d","n/d","n/d","n/d","Razoavel","Nativo","Basico+",
  "Muito rapido; DROP-IN; qualidade limitada p/ tarefas complexas","HF Qwen3"],

 ["SmolLM2 1.7B","HuggingFace","Densa",1.7,1.7,"8K","Apache-2.0",1.1,
  "Sim","~12-13 GB","~50-70",
  "n/d","n/d","n/d","n/d","Fraco (EN-centrico)","Limitado","Basico",
  "Leve, mas fraco em PT-BR e tools","HF HuggingFaceTB/SmolLM2"],

 ["Llama 3.2 1B Instruct","Meta","Densa",1.0,1.0,"128K","Llama 3.2 (custom)",0.8,
  "Sim","~13 GB","~60-90",
  "n/d",27.2,59.5,25.7,"Razoavel","Fraco (BFCL 26)","Basico",
  "Ultra-leve; tool calling fraco; so p/ comandos simples","HF meta-llama/Llama-3.2-1B"],

 ["Qwen3-0.6B","Qwen","Densa",0.6,0.6,"32K","Apache-2.0",0.5,
  "Sim","~13 GB","~80-100",
  "n/d","n/d","n/d","n/d","Fraco","Nativo (limitado)","Basico",
  "Minusculo; so p/ wake/roteamento, nao p/ conversa real","HF Qwen3"],

 # ---- Avaliados a pedido (jun/2026): MiniMax e Gemma 4 ----
 ["MiniMax-M1","MiniMax","Hibrido MoE (lightning attn)",456,45.9,"1M","Apache-2.0",250,
  "Nao","~0 (impossivel)","n/a",
  "n/d","n/d","n/d","n/d","Bom (multiling.)","Nativo (agentic)","Muito Alto",
  "Escala datacenter (456B totais): impossivel local; so servidor/API","arXiv 2506.13585"],

 ["MiniMax-M2","MiniMax","MoE (10B ativos)",229,10.0,"196K","MIT",130,
  "Nao","~0 (impossivel)","n/a",
  "n/d","n/d","n/d","n/d","Bom (multiling.)","Nativo (agentic forte)","Muito Alto",
  "229B totais: minimo ~64GB (IQ1_S). Impossivel neste hardware; servidor/API","HF unsloth/MiniMax-M2-GGUF"],

 ["Gemma 4 31B Dense","Google","Densa",31,31,"256K","Apache-2.0",19,
  "Nao","~0 livre","~2-3",
  "n/d","n/d","n/d","n/d (t2-bench 86.4)","Excelente (140 idiomas)","Nativo (FC)","Muito Alto",
  "Top de qualidade aberta (AIME 89.2), mas 31B denso pesa demais aqui","blog.google/gemma-4"],

 ["Gemma 4 26B-A4B (MoE)","Google","MoE (3.8B ativos)",26,3.8,"256K","Apache-2.0",16,
  "Nao","~0-2 GB","~3-5",
  "n/d","n/d","n/d","n/d","Excelente (140 idiomas)","Nativo (FC)","Alto",
  "MoE forte (AIME 88.3), mas 26B totais nao cabem (mesmo trap do 35B)","blog.google/gemma-4"],

 ["Gemma 4 12B","Google","Densa (multimodal)",12,12,"128K","Apache-2.0",7.3,
  "Parcial","~7-8 GB","~12-18",
  "n/d","n/d","n/d","n/d","Excelente (140 idiomas)","Nativo (FC)","Alto",
  "Multimodal forte; cabe quase 100% com ctx modesto; FC nativo","ai.google.dev/gemma"],

 ["Gemma 4 E4B","Google","Densa PLE (multimodal)",4.5,4.5,"128K","Apache-2.0",5.0,
  "Q4 so","~10-11 GB","~25-40",
  "n/d","n/d","n/d","n/d","Excelente (140 idiomas)","Nativo (FC)","Medio-Alto",
  "Rival direto do Qwen3-4B; PLE infla memoria (Q8 ~9-10GB nao cabe -> roda Q4); template/FC diferentes","ai.google.dev/gemma"],

 ["Gemma 4 E2B","Google","Densa PLE (multimodal)",2.3,2.3,"128K","Apache-2.0",2.9,
  "Sim","~11-12 GB","~40-60",
  "n/d","n/d","n/d","n/d","Excelente (140 idiomas)","Nativo (FC)","Medio",
  "Leve e multimodal; otimo multilingue; menos capaz que o E4B","ai.google.dev/gemma"],
]
# Mantem o modelo ATUAL no topo; ordena o resto por params totais (desc).
_base = [r for r in M if "ATUAL" in r[0]]
_rest = sorted([r for r in M if "ATUAL" not in r[0]], key=lambda r: -float(r[3]))
M = _base + _rest

HEADERS =["Modelo","Dev","Arquitetura","Params\nTotais (B)","Params\nAtivos (B)","Contexto",
 "Licenca","GGUF\nQ4_K_M (GB)","Cabe 100%\nVRAM 8GB?","RAM livre\nest. (16GB)","Veloc. est.\nRX580 (tok/s)",
 "MMLU-Pro","GPQA","IFEval","BFCL\n(tool)","PT-BR /\nMultilingue","Tool\ncalling","Qualidade\n(tier)",
 "Veredito p/ assistente de voz local","Fonte\n(benchmarks)"]

wb = Workbook()

# =========================================================================
# ABA 1 — Comparacao
# =========================================================================
ws = wb.active
ws.title = "Comparacao"

# Titulo
ws.merge_cells("A1:T1")
t = ws["A1"]
t.value = "Jarvis-1 — Comparativo de Modelos LLM locais  (hardware: Ryzen 5 4500 - 16GB RAM - RX 580 8GB)"
t.font = Font(name=FONT, size=13, bold=True, color="FFFFFF")
t.fill = PatternFill("solid", fgColor=C_HDR)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 26

ws.merge_cells("A2:T2")
s = ws["A2"]
s.value = ("Veredito por EQUILIBRIO (footprint x velocidade x qualidade).  "
           "* = recomendado.  Benchmarks: fontes oficiais (ver aba Fontes); 'n/d' = nao coletado de fonte primaria.  "
           "Footprint/velocidade = estimativa para este hardware.")
s.font = Font(name=FONT, size=9, italic=True, color="404040")
s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 28

# Header row (row 3)
hr = 3
for c, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=C_HDR)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
ws.row_dimensions[hr].height = 40

# Data
def fits_fill(v):
    if v == "Sim": return C_GOOD, C_GOODT
    if v == "Parcial": return C_PART, "7F6000"
    return "F8CBAD", C_BADT

r = hr + 1
for i, row in enumerate(M):
    is_base = "ATUAL" in row[0]
    is_rec  = "*" in row[0]
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(name=FONT, size=9,
                         bold=(c == 1),
                         color="FFFFFF" if is_base and c == 1 else "000000")
        cell.alignment = Alignment(horizontal="left" if c in (1,3,6,7,10,19,20) else "center",
                                   vertical="center", wrap_text=(c in (1,19)))
        cell.border = BORDER
        # zebra
        if not is_base and i % 2 == 1 and c != 9:
            cell.fill = PatternFill("solid", fgColor=C_ALT)
    # destaque coluna "Cabe VRAM" (col 9)
    fcell = ws.cell(row=r, column=9)
    ffill, ftxt = fits_fill(row[8])
    fcell.fill = PatternFill("solid", fgColor=ffill)
    fcell.font = Font(name=FONT, size=9, bold=True, color=ftxt)
    # destaque modelo
    if is_base:
        for c in range(1, 21):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=C_BASE)
            cc = ws.cell(row=r, column=c)
            cc.font = Font(name=FONT, size=9, bold=(c in (1,)), color="FFFFFF")
        ws.cell(row=r, column=9).font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
    if is_rec:
        vcell = ws.cell(row=r, column=19)
        vcell.fill = PatternFill("solid", fgColor=C_GOOD)
        vcell.font = Font(name=FONT, size=9, bold=True, color=C_GOODT)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=C_GOOD)
        ws.cell(row=r, column=1).font = Font(name=FONT, size=9, bold=True, color=C_GOODT)
    ws.row_dimensions[r].height = 42
    r += 1

# Larguras
widths = [30,11,26,9,9,12,17,11,11,13,12,9,7,8,8,16,16,12,46,24]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "C4"
ws.auto_filter.ref = f"A{hr}:T{r-1}"

# =========================================================================
# ABA 2 — Analise & Recomendacao
# =========================================================================
wa = wb.create_sheet("Analise e Recomendacao")
wa.column_dimensions["A"].width = 3
wa.column_dimensions["B"].width = 110

def block(ws_, row, title, body, title_color=C_HDR, body_size=10):
    c = ws_.cell(row=row, column=2, value=title)
    c.font = Font(name=FONT, size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=title_color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_.row_dimensions[row].height = 24
    row += 1
    for para in body:
        cc = ws_.cell(row=row, column=2, value=para)
        cc.font = Font(name=FONT, size=body_size, color="000000")
        cc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        # altura proporcional ao texto
        ws_.row_dimensions[row].height = max(16, 16 * (len(para)//95 + 1))
        row += 1
    row += 1
    return row

rr = 1
ttl = wa.cell(row=rr, column=2, value="Analise: qual modelo substitui o 35B-A3B mantendo a maquina utilizavel?")
ttl.font = Font(name=FONT, size=14, bold=True, color=C_HDR)
rr += 2

rr = block(wa, rr, "1. O problema do modelo atual (Qwen3.6-35B-A3B)", [
 "- E um modelo Qwen oficial, geracao 3.6, arquitetura hibrida (Gated DeltaNet + atencao 3:1) com foco em coding agentico.",
 "- Tem 35B de parametros TOTAIS mas so 3B ATIVOS por token (MoE). A qualidade de raciocinio por token fica na classe de um modelo denso ~3-4B; o ganho do MoE e a amplitude de conhecimento (35B).",
 "- POReM: paga o custo de MEMORIA dos 35B totais. No quant Q4_K_M sao 19.7 GB. Como nao cabe na VRAM (8GB) nem na RAM (16GB) de forma confortavel, ele usa mmap e ocupa praticamente TODA a RAM+VRAM -> nada mais roda na maquina.",
 "- Velocidade medida neste hardware: ~5-8 tok/s, com ~100s de carga no startup. Funciona, mas inviabiliza usar o PC para outras coisas (o oposto do proposito de um assistente local).",
])

rr = block(wa, rr, "2. Achado-chave (fonte oficial Qwen)", [
 "O card oficial do Qwen3-4B-Instruct-2507 mostra que ele EMPATA OU SUPERA o Qwen3-30B-A3B (3B ativos, irmao do modelo atual) na maioria dos benchmarks:",
 "   - MMLU-Pro: 69.6 (4B) vs 69.1 (30B-A3B)      - GPQA: 62.0 vs 54.8      - AIME25: 47.4 vs 21.6",
 "   - Arena-Hard v2: 43.4 vs 24.8                 - BFCL (tool calling): 61.9 vs 58.6",
 "   - Unicos onde o 30B-A3B ganha por pouco: IFEval (83.7 vs 83.4) e MultiIF multilingue (70.8 vs 69.0).",
 "Ou seja: um modelo denso moderno de 4B entrega qualidade equivalente a do A3B usando ~1/8 da memoria (2.5 GB vs 18-20 GB).",
], title_color="006100")

rr = block(wa, rr, "3. Recomendacao por perfil (equilibrio)", [
 ">> MELHOR ESCOLHA GERAL:  Qwen3-4B-Instruct-2507",
 "   - Voce JA baixou (esta em A:). DROP-IN no codigo atual (mesmo chat template, mesmo --reasoning off, mesmas tools Qwen).",
 "   - 2.5 GB cabem 100% na VRAM do RX 580 -> a RAM (16GB) fica quase toda livre p/ outros apps. Estimado ~30-45 tok/s.",
 "   - Qualidade equivalente ao 30B-A3B; 262K de contexto; tool calling nativo (BFCL 61.9).",
 "",
 ">> SE QUISER MAIS 'MUSCULO' E ACEITAR menos folga:  Qwen3-8B (tambem DROP-IN, ja usado no projeto)",
 "   - 4.9 GB ainda cabem na VRAM (com contexto modesto); deixa ~9-10 GB de RAM livre; ~18-28 tok/s.",
 "",
 ">> SE PT-BR / MULTILINGUE FOR PRIORIDADE:  Gemma 3 4B  (melhor multilingue leve, 140 idiomas, 128K ctx)",
 "   - Cuidado: licenca Gemma (restritiva) e tool calling menos maduro -> exigiria ajuste no llm_local.py.",
 "",
 ">> MAXIMA VELOCIDADE / HARDWARE MINIMO:  Qwen3-1.7B  (DROP-IN, ~50-70 tok/s) — qualidade basica, bom p/ comandos simples.",
], title_color="006100")

rr = block(wa, rr, "4. Nota de migracao (esforco no codigo)", [
 "- Familia Qwen (4B-2507, 8B, 1.7B, 30B-A3B): troca trivial — so mudar LLAMA_MODEL_PATH/QWEN_MODEL no .env. Mesmo template, mesmo --reasoning off, mesma amostragem (top_k/top_p/presence_penalty), mesmas tools.",
 "- Llama 3.x / Gemma 3 / Phi-4 / Mistral: funcionam no llama-server, MAS usam outro chat template e outro formato de tool calling -> precisa ajustar llm_local.py (template e parsing de tool_calls). Esforco medio.",
 "- Modelos densos <=8B cabem 100% na VRAM e NAO precisam do build TurboQuant nem do --n-cpu-moe (aquilo so era necessario por causa do MoE gigante). Pode-se usar ate o llama-server padrao.",
])

rr = block(wa, rr, "4b. Avaliados a pedido: Gemma 4 e MiniMax (jun/2026)", [
 "MiniMax (M1/M2/M2.1): FORA por hardware, nao por qualidade. Sao modelos de escala datacenter (229B-456B totais). Mesmo com 10B ativos, os totais tem que caber na memoria: o menor GGUF do M2 (IQ1_S, 1-bit) ja sao 64 GB. No RX 580 (8GB+16GB) e fisicamente impossivel -> so vLLM/SGLang em servidor ou API. E o mesmo erro do 35B, 3-13x pior.",
 "Gemma 4 (mar/2026): virou concorrente legitimo (agora Apache-2.0, function calling nativo, GGUF drop-in). O E4B e o rival direto do Qwen3-4B. Mesmo assim o Qwen vence NESTE caso por 3 motivos: (1) Q8 cabe nos 8GB de VRAM, enquanto o E4B usa Per-Layer Embeddings que inflam a memoria (Q8 ~9-10GB nao cabe -> roda Q4, qualidade menor); (2) Qwen e drop-in no codigo atual (Gemma exige outro chat template e outro formato de tool calling); (3) Qwen3-4B denso e maduro no llama.cpp/Vulkan, enquanto o PLE do Gemma 4 e arquitetura nova (risco de repetir a dor bleeding-edge do 35B).",
 "Onde o Gemma 4 PODE ganhar: PT-BR/multilingue (140 idiomas, tradicionalmente forte). Se a fluencia em portugues virar prioridade #1, vale testar o E4B aceitando Q4 + mexida no codigo.",
], title_color="7F6000")

rr = block(wa, rr, "5. Conclusao", [
 "Para um assistente de VOZ local que deve COEXISTIR com o resto da maquina, o 35B-A3B e exagero: gasta 8x mais memoria para uma qualidade que um Qwen3-4B-Instruct-2507 entrega cabendo inteiro na VRAM.",
 "Recomendacao: migrar para Qwen3-4B-Instruct-2507 (ja baixado, drop-in). Manter o 35B disponivel apenas para tarefas pontuais de coding agentico que justifiquem ocupar a maquina toda.",
], title_color="C00000")

# =========================================================================
# ABA 3 — Fontes & Metodologia
# =========================================================================
wf = wb.create_sheet("Fontes e Metodologia")
wf.column_dimensions["A"].width = 3
wf.column_dimensions["B"].width = 50
wf.column_dimensions["C"].width = 75

hf = wf.cell(row=1, column=2, value="Fontes (coletadas em jun/2026)")
hf.font = Font(name=FONT, size=13, bold=True, color=C_HDR)
wf.cell(row=2, column=2, value="O que").font = Font(name=FONT, bold=True, color="FFFFFF")
wf.cell(row=2, column=3, value="URL / Referencia").font = Font(name=FONT, bold=True, color="FFFFFF")
for c in (2,3):
    wf.cell(row=2, column=c).fill = PatternFill("solid", fgColor=C_HDR)

FONTES = [
 ["Qwen3-4B-Instruct-2507 (benchmarks oficiais)","https://huggingface.co/Qwen/Qwen3-4B-Instruct-2507"],
 ["Qwen3.6-35B-A3B (arquitetura / modelo atual)","https://qwen.ai/blog?id=qwen3.6-35b-a3b  |  https://huggingface.co/HauhauCS/Qwen3.6-35B-A3B-Uncensored-HauhauCS-Aggressive"],
 ["Qwen3 Technical Report (familia 0.6B-235B)","https://arxiv.org/abs/2505.09388"],
 ["Qwen3 — guia de specs por tamanho","https://insiderllm.com/guides/qwen3-complete-guide/"],
 ["Gemma 3 (1B/4B/12B/27B)","https://huggingface.co/google/gemma-3-4b-it"],
 ["Phi-4-mini-instruct (3.8B)","https://huggingface.co/microsoft/Phi-4-mini-instruct"],
 ["Llama 3.2 (1B/3B) + Llama 3.1 8B","https://huggingface.co/meta-llama/Llama-3.2-3B-Instruct  |  arXiv:2407.21783"],
 ["Ministral 3 (8B) paper","https://arxiv.org/abs/2601.08584"],
 ["Llama 3.1 8B — analise comparativa","https://artificialanalysis.ai/models/llama-3-1-instruct-8b"],
 ["Gemma 4 (anuncio oficial)","https://blog.google/innovation-and-ai/technology/developers-tools/gemma-4/"],
 ["Gemma 4 (specs E2B/E4B, PLE, function calling, licenca)","https://ai.google.dev/gemma/docs/core"],
 ["MiniMax-M2 GGUF (tamanhos / VRAM minima)","https://huggingface.co/unsloth/MiniMax-M2-GGUF"],
 ["MiniMax-M1 (456B, lightning attention) paper","https://arxiv.org/abs/2506.13585"],
]
rf = 3
for nome, url in FONTES:
    wf.cell(row=rf, column=2, value=nome).font = Font(name=FONT, size=10)
    wf.cell(row=rf, column=3, value=url).font = Font(name=FONT, size=9, color="0563C1")
    for c in (2,3):
        wf.cell(row=rf, column=c).alignment = Alignment(vertical="center", wrap_text=True)
        wf.cell(row=rf, column=c).border = BORDER
    rf += 1

rf += 1
met = wf.cell(row=rf, column=2, value="Metodologia das colunas de hardware")
met.font = Font(name=FONT, size=12, bold=True, color=C_HDR); rf += 1
NOTAS = [
 "GGUF Q4_K_M (GB): tamanho do arquivo no quant Q4_K_M (~4.8 bits/peso). 35B medido = 19.7 GB; demais estimados por params.",
 "Cabe 100% VRAM 8GB?: 'Sim' se modelo+KV+buffers <= ~7 GB (sobra p/ contexto). RX 580 util ~7.4 GB.",
 "RAM livre est.: de 16 GB, descontando Windows (~3-4 GB) + venv/Whisper/Piper (~2-3 GB). Se o modelo cabe na VRAM, a RAM fica livre.",
 "Veloc. est. RX580 (tok/s): estimativa memory-bound (RX 580 ~256 GB/s). Modelos na VRAM: rapidos. Com offload p/ CPU: lentos. 35B medido ~5-8 tok/s.",
 "Benchmarks 'n/d': nao coletados de fonte primaria nesta pesquisa — evitei numeros sem fonte. Colunas Qualidade/PT-BR/Tools sao avaliacao qualitativa do analista.",
 "Benchmarks nao sao 100% comparaveis entre familias (setups/harness diferentes; thinking vs non-thinking). Use como ordem de grandeza.",
]
for n in NOTAS:
    cc = wf.cell(row=rf, column=2, value=n)
    cc.font = Font(name=FONT, size=9, color="000000")
    cc.alignment = Alignment(wrap_text=True, vertical="top")
    wf.merge_cells(start_row=rf, start_column=2, end_row=rf, end_column=3)
    wf.row_dimensions[rf].height = max(16, 16*(len(n)//120 + 1))
    rf += 1

import os
out = r"C:\Users\Usuario\VSCodeProjects\jarvis-1\docs\comparativo_modelos_LLM.xlsx"
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("OK ->", out)
print("Modelos:", len(M))
